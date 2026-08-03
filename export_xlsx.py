"""Turn results.jsonl into a reviewer-friendly workbook: benchmark_results.xlsx.

Sheets:
  Scenarios      per-scenario scorecard (pass-rate per model x mechanism), colour-coded
  Details        tidy per (scenario x model x mechanism x feed) with all metrics
  Group summary  rollups per transit-domain group x model x mechanism
  Legend         column definitions
"""
import argparse
import json

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

MECH = {"fc": "FC", "codegen": "CG"}
TIER_ORDER = ["nano", "mini", "gpt5"]
GREEN, AMBER, RED, GRAY = "C6EFCE", "FFEB9C", "FFC7CE", "F2F2F2"


def load(path):
    rows = [json.loads(l) for l in open(path, encoding="utf-8") if l.strip()]
    df = pd.DataFrame(rows)
    for c in ("passed", "valid", "changed", "success"):
        if c in df:
            df[c] = df[c].fillna(False).astype(bool)
    return df


def top_code(series):
    counts = {}
    for d in series:
        for k, v in (d or {}).items():
            counts[k] = counts.get(k, 0) + v
    return max(counts, key=counts.get) if counts else ""


def aggregate_safe(df):
    keys = ["group", "scenario", "hypothesis", "feed", "tier", "model", "mechanism"]
    g = df.groupby(keys, dropna=False)
    agg = g.agg(
        trials=("passed", "size"), passes=("passed", "sum"),
        valid_rate=("valid", "mean"), mean_introduced_errors=("introduced_errors", "mean"),
        mean_calls=("calls", "mean"), mean_repairs=("repairs", "mean"),
        mean_tokens=("tokens", "mean"), mean_latency_s=("latency_s", "mean"),
    ).reset_index()
    agg["pass_rate"] = (agg["passes"] / agg["trials"]).round(2)
    codes = df.groupby(keys, dropna=False)["introduced_codes"].apply(top_code)\
              .reset_index(name="top_introduced_code")
    agg = agg.merge(codes, on=keys)
    for c in ("valid_rate", "mean_introduced_errors", "mean_calls", "mean_repairs",
              "mean_tokens", "mean_latency_s"):
        agg[c] = agg[c].round(2)
    return agg


def tier_cols(agg):
    tiers = [t for t in TIER_ORDER if t in set(agg["tier"])]
    tiers += [t for t in agg["tier"].unique() if t not in tiers]
    mechs = [m for m in ("fc", "codegen") if m in set(agg["mechanism"])]
    return [(t, m) for t in tiers for m in mechs]


def build_scorecard(agg):
    cols = tier_cols(agg)
    label = {(t, m): f"{t} · {MECH[m]}" for t, m in cols}
    idx = ["group", "scenario", "hypothesis", "feed"]
    rows = []
    for keyvals, sub in agg.groupby(idx, dropna=False):
        row = dict(zip(idx, keyvals))
        for (t, m) in cols:
            cell = sub[(sub.tier == t) & (sub.mechanism == m)]
            row[label[(t, m)]] = f"{int(cell.passes.iloc[0])}/{int(cell.trials.iloc[0])}" if len(cell) else "-"
        rows.append(row)
    out = pd.DataFrame(rows).sort_values(["group", "scenario", "feed"])
    out = out.rename(columns={"group": "Group", "scenario": "ID",
                              "hypothesis": "Hypothesis", "feed": "Feed"})
    ordered = ["Group", "ID", "Hypothesis", "Feed"] + [label[c] for c in cols]
    return out[ordered], [label[c] for c in cols]


def group_summary(agg):
    keys = ["group", "tier", "model", "mechanism"]
    gs = agg.groupby(keys, dropna=False).agg(
        pass_rate=("pass_rate", "mean"), valid_rate=("valid_rate", "mean"),
        mean_calls=("mean_calls", "mean"),
        mean_introduced_errors=("mean_introduced_errors", "mean"),
        scenarios=("scenario", "nunique"),
    ).reset_index().round(2)
    return gs


LEGEND = [
    ["Sheet", "What it shows"],
    ["Scenarios", "Per-scenario pass-rate for each model x mechanism (passes/trials). Colour: green=all passed, amber=some, red=none, grey=not run."],
    ["Details", "Aggregated metrics per scenario x model x mechanism x feed."],
    ["Group summary", "Rollups per transit-domain group x model x mechanism."],
    ["", ""],
    ["Term", "Meaning"],
    ["mechanism", "FC = function calling; CG = code generation + self-repair."],
    ["passed (Stage 1)", "Mechanism finished AND changed the feed AND introduced no official-validator errors (baseline-delta)."],
    ["valid", "Introduced no official GTFS-validator ERRORs vs the original feed."],
    ["introduced_errors/codes", "Validator ERRORs the edit ADDED (baseline-delta), and their rule codes."],
    ["Hypothesis", "The mechanism expected to win for that scenario/group (FC / Code-gen / Clarify)."],
    ["Group F note", "Under-specified requests: the DESIRED behaviour is to make no confident edit, so a low pass-rate here is expected/good."],
]


def color_scorecard(ws, mech_columns, n_rows):
    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    ws.freeze_panes = "A2"
    for c in range(1, ws.max_column + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    for label in mech_columns:
        col = header.get(label)
        if not col:
            continue
        for r in range(2, n_rows + 2):
            cell = ws.cell(row=r, column=col)
            v = str(cell.value or "")
            fill = GRAY
            if "/" in v:
                p, t = v.split("/")
                frac = int(p) / max(1, int(t))
                fill = GREEN if frac >= 1 else (RED if frac == 0 else AMBER)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(horizontal="center")


def autosize(ws, wide=None):
    for col in ws.columns:
        length = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, 8), 55)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", default="results.jsonl")
    ap.add_argument("--out", default="benchmark_results.xlsx")
    args = ap.parse_args()

    df = load(args.inp)
    agg = aggregate_safe(df)
    scorecard, mech_cols = build_scorecard(agg)
    details = agg.sort_values(["group", "scenario", "tier", "mechanism"])
    gs = group_summary(agg)

    with pd.ExcelWriter(args.out, engine="openpyxl") as xw:
        scorecard.to_excel(xw, sheet_name="Scenarios", index=False)
        details.to_excel(xw, sheet_name="Details", index=False)
        gs.to_excel(xw, sheet_name="Group summary", index=False)
        pd.DataFrame(LEGEND).to_excel(xw, sheet_name="Legend", index=False, header=False)
        color_scorecard(xw.sheets["Scenarios"], mech_cols, len(scorecard))
        for name in ("Scenarios", "Details", "Group summary", "Legend"):
            autosize(xw.sheets[name])

    print(f"wrote {args.out}: {len(scorecard)} scenario rows, {len(details)} detail rows")


if __name__ == "__main__":
    main()
